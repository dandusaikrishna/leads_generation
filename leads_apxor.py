"""
BACKWARDS COMPATIBLE WRAPPER - Legacy Leads Pipeline

This module maintains compatibility with the original leads_apxor.py while
delegating to the new modular pipeline architecture.

NEW DEVELOPMENT: Please use python leads.py or from pipeline import DiscoveryPipeline

This script is preserved to ensure existing automation continues to work.
"""

import os
import sys
import logging
import time
import re
import socket
import smtplib
import random
import string
import json
import dns.resolver
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import new modular components
from utils import setup_logger
from config import SERPER_API_KEY, OPENROUTER_API_KEY
from pipeline import DiscoveryPipeline
from exporters import export_companies_json, export_emails_txt

# Setup legacy logger
log = setup_logger("enrich_leads_hiring_legacy")


# ─────────────────────────────────────────────────────────────────────────────
# Legacy Wrapper Functions - Backwards Compatibility
# ─────────────────────────────────────────────────────────────────────────────

def discover_hiring_companies(roles: list, country: str, limit: int) -> list:
    """Legacy wrapper - forwards to new pipeline"""
    log.info(f"Using legacy API - forwarding to new pipeline")
    from company_discovery import discover_hiring_companies as new_discover
    return new_discover(roles, country, limit)


def resolve_company_domain(company_name: str, country: str) -> str:
    """Legacy wrapper"""
    from company_info import resolve_company_domain as new_resolve
    result = new_resolve(company_name, country)
    return result or ""


def extract_decision_makers(company_name: str, domain: str, country: str) -> list:
    """Legacy wrapper"""
    from contact_discovery import find_senior_contacts
    return find_senior_contacts(company_name, domain, country)


def split_name(fullname: str) -> tuple:
    """Legacy wrapper"""
    from utils import split_name as new_split
    return new_split(fullname)


def generate_email_candidates(fullname: str, domain: str) -> list:
    """Legacy wrapper - returns email strings"""
    from email_patterns import generate_email_candidates as new_generate
    candidates = new_generate(fullname, domain)
    return [c["email"] for c in candidates]


# ─────────────────────────────────────────────────────────────────────────────
# SMTP Email Verification (Legacy - Preserved for compatibility)
# ─────────────────────────────────────────────────────────────────────────────

MX_CACHE = {}
CATCH_ALL_CACHE = {}
DOMAINS_FAILED_CONN = set()


def get_mx_hosts(domain: str, timeout: float = 8.0) -> list:
    domain = domain.strip().lower()
    if domain in MX_CACHE:
        return MX_CACHE[domain]
    try:
        answers = dns.resolver.resolve(domain, 'MX')
        hosts = [str(r.exchange).rstrip('.').lower() for r in sorted(answers, key=lambda r: r.preference)]
        MX_CACHE[domain] = hosts
        return hosts
    except Exception as e:
        log.debug(f"MX lookup failed for {domain}: {e}")
        MX_CACHE[domain] = []
        return []


def check_catch_all(domain: str, mx_hosts: list, timeout: float = 6.0) -> bool:
    domain = domain.strip().lower()
    if domain in CATCH_ALL_CACHE:
        return CATCH_ALL_CACHE[domain]
    if not mx_hosts or domain in DOMAINS_FAILED_CONN:
        return False
    
    rand_prefix = "".join(random.choice(string.ascii_lowercase + string.digits) for _ in range(16))
    bogus_email = f"verify-test-{rand_prefix}@{domain}"
    
    log.info(f"Checking if domain '{domain}' is catch-all...")
    
    connected_any = False
    for mx in mx_hosts[:2]:
        try:
            with smtplib.SMTP(mx, 25, timeout=timeout) as smtp:
                connected_any = True
                smtp.helo("deepta.ai")
                smtp.mail("verify@deepta.ai")
                code, msg = smtp.rcpt(bogus_email)
                if code in (250, 251, 252):
                    CATCH_ALL_CACHE[domain] = True
                    return True
                elif code >= 500:
                    CATCH_ALL_CACHE[domain] = False
                    return False
        except Exception as e:
            log.debug(f"SMTP failed on {mx}: {e}")
    
    if not connected_any:
        DOMAINS_FAILED_CONN.add(domain)
    return False


def verify_email_smtp(email: str, domain: str, mx_hosts: list, timeout: float = 6.0) -> tuple:
    if not mx_hosts:
        return "Invalid", 0, "No MX records found"
    if domain in DOMAINS_FAILED_CONN:
        return "Unknown", 30, "MX servers unreachable (cached)"
    
    is_catch_all = check_catch_all(domain, mx_hosts, timeout)
    if domain in DOMAINS_FAILED_CONN:
        return "Unknown", 30, "MX servers unreachable"
    
    connected_any = False
    for mx in mx_hosts[:2]:
        try:
            with smtplib.SMTP(mx, 25, timeout=timeout) as smtp:
                connected_any = True
                smtp.helo("deepta.ai")
                smtp.mail("verify@deepta.ai")
                code, msg = smtp.rcpt(email)
                msg_text = msg.decode("utf-8", errors="ignore") if isinstance(msg, bytes) else str(msg)
                
                block_keywords = ["spam", "block", "blacklist", "denied", "abuse", "reputation"]
                is_blocked = any(kw in msg_text.lower() for kw in block_keywords)
                
                if code in (250, 251, 252):
                    if is_blocked:
                        return "Unknown", 40, f"Accepted but blocked keywords at {mx}"
                    if is_catch_all:
                        return "Catch-all / Probable", 65, f"Catch-all domain"
                    else:
                        return "Verified Valid", 95, f"Verified at {mx}"
                elif code >= 500:
                    return "Invalid", 10, f"Hard reject at {mx}"
                elif code >= 400:
                    return "Unknown", 35, f"Temp reject at {mx}"
        except socket.timeout:
            continue
        except Exception as e:
            log.debug(f"SMTP error on {mx}: {e}")
            continue
    
    if not connected_any:
        DOMAINS_FAILED_CONN.add(domain)
    return "Unknown", 30, "All connections failed"


# ─────────────────────────────────────────────────────────────────────────────
# Excel Writer (Legacy)
# ─────────────────────────────────────────────────────────────────────────────

def write_styled_excel(data_rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hiring Leads"
    
    headers = [
        "Company Name", "Official Website", "Hiring Role", "Job Post URL",
        "Lead Name", "Lead Designation", "LinkedIn URL", "Email ID", 
        "Email Status", "Verification Score", "Verification Notes"
    ]
    
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    hdr_font = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
    
    status_colors = {
        "Verified Valid": PatternFill("solid", fgColor="C8E6C9"),
        "Catch-all / Probable": PatternFill("solid", fgColor="FFF9C4"),
        "Unknown": PatternFill("solid", fgColor="F5F5F5"),
        "Invalid": PatternFill("solid", fgColor="FFCDD2"),
        "Invalid/No Email Found": PatternFill("solid", fgColor="FFCDD2"),
        "No Contacts Found": PatternFill("solid", fgColor="EEEEEE"),
    }
    
    font_regular = Font(size=10, name="Segoe UI")
    font_bold = Font(bold=True, size=10, name="Segoe UI")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    for r_idx, row_data in enumerate(data_rows, 2):
        for col_idx, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            if h in ("Official Website", "Job Post URL", "LinkedIn URL") and val and val != "N/A":
                cell.font = Font(color="0000EE", underline="single", size=10, name="Segoe UI")
            if h == "Verification Score":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if h == "Email Status":
                cell.fill = status_colors.get(val, status_colors.get("Unknown"))
                cell.font = font_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    col_widths = {1: 22, 2: 24, 3: 25, 4: 30, 5: 20, 6: 26, 7: 35, 8: 28, 9: 22, 10: 18, 11: 45}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 28
    for r in range(2, len(data_rows) + 2):
        ws.row_dimensions[r].height = 22
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main Pipeline - Unified Approach
# ─────────────────────────────────────────────────────────────────────────────

def run_hiring_enrichment_pipeline(roles: list, country: str, limit: int, output_excel_path: str, cache_file: str):
    """Main entry point - delegates to new pipeline"""
    log.info("Starting legacy pipeline wrapper...")
    log.info(f"Country: {country}, Roles: {roles}, Limit: {limit}")
    
    pipeline = DiscoveryPipeline()
    companies = pipeline.discover_and_process_companies(
        country=country,
        roles=roles,
        limit=limit,
        parallel=True
    )
    
    if not companies:
        log.error("No companies discovered")
        return
    
    log.info(f"Exporting {len(companies)} companies...")
    export_companies_json(companies, "leads/companies.json")
    export_emails_txt(companies, "leads/emails.txt")
    
    if output_excel_path:
        log.info(f"Generating Excel report: {output_excel_path}")
        excel_data = []
        for company in companies:
            if not company.contacts:
                excel_data.append({
                    "Company Name": company.company_name,
                    "Official Website": company.website or "",
                    "Hiring Role": company.jobs[0].title if company.jobs else "N/A",
                    "Job Post URL": company.jobs[0].url if company.jobs else "N/A",
                    "Lead Name": "N/A",
                    "Lead Designation": "N/A",
                    "LinkedIn URL": "N/A",
                    "Email ID": "N/A",
                    "Email Status": "No Contacts Found",
                    "Verification Score": 0,
                    "Verification Notes": "No contacts"
                })
            else:
                for contact in company.contacts[:5]:
                    email = company.public_emails[0].email if company.public_emails else ""
                    excel_data.append({
                        "Company Name": company.company_name,
                        "Official Website": company.website or "",
                        "Hiring Role": company.jobs[0].title if company.jobs else "N/A",
                        "Job Post URL": company.jobs[0].url if company.jobs else "N/A",
                        "Lead Name": contact.name,
                        "Lead Designation": contact.role,
                        "LinkedIn URL": contact.public_profile_url or "N/A",
                        "Email ID": email or "N/A",
                        "Email Status": "Probable" if email else "N/A",
                        "Verification Score": contact.role_score,
                        "Verification Notes": f"Score: {contact.role_score}"
                    })
        
        write_styled_excel(excel_data, output_excel_path)
        log.info(f"Excel report saved: {output_excel_path}")
    
    log.info("Pipeline complete!")


# ─────────────────────────────────────────────────────────────────────────────
# CLI Entry Point (Legacy)
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="[LEGACY] Find companies currently hiring and enrich leads.",
        epilog="This is the legacy script. For new projects, use: python leads.py"
    )
    parser.add_argument("--country", default="USA", help="Target country")
    parser.add_argument("--roles", default="Software Engineer,Backend Engineer", help="Comma-separated roles")
    parser.add_argument("--limit", type=int, default=50, help="Max companies to discover")
    parser.add_argument("--output", default="leads/hiring_leads_enriched.xlsx", help="Output Excel file")
    parser.add_argument("--cache", default="leads_hiring_cache.json", help="Cache file path")
    
    args = parser.parse_args()
    roles_list = [r.strip() for r in args.roles.split(",") if r.strip()]
    
    log.info("=" * 70)
    log.info("LEGACY API MODE - Maintained for Backwards Compatibility")
    log.info("For new projects, use: python leads.py")
    log.info("=" * 70)
    
    try:
        run_hiring_enrichment_pipeline(
            roles=roles_list,
            country=args.country,
            limit=args.limit,
            output_excel_path=args.output,
            cache_file=args.cache
        )
        log.info("=" * 70)
        log.info("Pipeline completed successfully!")
        log.info("=" * 70)
        sys.exit(0)
    except KeyboardInterrupt:
        log.warning("Pipeline interrupted")
        sys.exit(130)
    except Exception as e:
        log.error(f"Pipeline error: {e}", exc_info=True)
        sys.exit(1)
