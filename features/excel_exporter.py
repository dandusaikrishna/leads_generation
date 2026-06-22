"""
Excel Export Module - Create styled Excel reports
"""

import os
from datetime import datetime
from typing import List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_leads_to_excel(leads: List[Dict], output_path: str = "output/leads.xlsx"):
    """
    Export leads to a styled Excel file.
    
    Args:
        leads: List of lead dictionaries
        output_path: Output file path
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Define headers based on lead structure
    headers = [
        "Company Name", "Website", "Industry", "Location",
        "Lead Name", "Lead Role", "Lead Email", "Email Status",
        "Phone", "LinkedIn URL", "Email Confidence",
        "Discovered Date", "Notes"
    ]
    
    # Header styling
    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
    
    # Status color mapping
    status_colors = {
        "Verified": "C8E6C9",      # Light green
        "Probable": "FFF9C4",      # Light yellow
        "Unknown": "F5F5F5",       # Light gray
        "Invalid": "FFCDD2",       # Light red
    }
    
    # Cell styles
    font_regular = Font(size=10, name="Segoe UI")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Write data rows
    for row_idx, lead in enumerate(leads, 2):
        # Company Name
        ws.cell(row=row_idx, column=1, value=lead.get("company_name", ""))
        # Website
        ws.cell(row=row_idx, column=2, value=lead.get("website", ""))
        # Industry
        ws.cell(row=row_idx, column=3, value=lead.get("industry", ""))
        # Location
        ws.cell(row=row_idx, column=4, value=lead.get("location", ""))
        # Lead Name
        ws.cell(row=row_idx, column=5, value=lead.get("lead_name", ""))
        # Lead Role
        ws.cell(row=row_idx, column=6, value=lead.get("lead_role", ""))
        # Lead Email
        ws.cell(row=row_idx, column=7, value=lead.get("email", ""))
        # Email Status
        status = lead.get("email_status", "Unknown")
        ws.cell(row=row_idx, column=8, value=status)
        # Phone
        ws.cell(row=row_idx, column=9, value=lead.get("phone", ""))
        # LinkedIn URL
        ws.cell(row=row_idx, column=10, value=lead.get("linkedin_url", ""))
        # Email Confidence
        conf = lead.get("email_confidence", 0)
        ws.cell(row=row_idx, column=11, value=conf)
        # Discovered Date
        ws.cell(row=row_idx, column=12, value=lead.get("discovered_date", ""))
        # Notes
        ws.cell(row=row_idx, column=13, value=lead.get("notes", ""))
        
        # Apply colors to status cells
        status_cell = ws.cell(row=row_idx, column=8)
        if status in status_colors:
            status_cell.fill = PatternFill(start_color=status_colors[status], 
                                         end_color=status_colors[status], 
                                         fill_type="solid")
        
        # Apply borders and font to all cells in row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.font = font_regular
            if col not in [2, 10]:  # Not URL columns
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Column widths
    col_widths = {
        1: 20,   # Company Name
        2: 25,   # Website
        3: 15,   # Industry
        4: 20,   # Location
        5: 18,   # Lead Name
        6: 20,   # Lead Role
        7: 25,   # Lead Email
        8: 15,   # Email Status
        9: 15,   # Phone
        10: 30,  # LinkedIn URL
        11: 15,  # Email Confidence
        12: 18,  # Discovered Date
        13: 30,  # Notes
    }
    
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add autofilter
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(output_path)
    return output_path


def export_emails_to_excel(emails: List[Dict], output_path: str = "output/emails.xlsx"):
    """Export emails to Excel with scoring."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    
    headers = ["Email", "Company", "Role", "Type", "Score", "Quality", "Website"]
    
    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    quality_colors = {
        "HIGH": "C8E6C9",
        "MEDIUM": "FFF9C4",
        "LOW": "FFCDD2",
    }
    
    for row_idx, email_data in enumerate(emails, 2):
        ws.cell(row=row_idx, column=1, value=email_data.get("email", ""))
        ws.cell(row=row_idx, column=2, value=email_data.get("company", ""))
        ws.cell(row=row_idx, column=3, value=email_data.get("role", ""))
        ws.cell(row=row_idx, column=4, value=email_data.get("type", ""))
        ws.cell(row=row_idx, column=5, value=email_data.get("score", 0))
        
        quality = email_data.get("quality", "LOW")
        quality_cell = ws.cell(row=row_idx, column=6, value=quality)
        if quality in quality_colors:
            quality_cell.fill = PatternFill(start_color=quality_colors[quality],
                                          end_color=quality_colors[quality],
                                          fill_type="solid")
        
        ws.cell(row=row_idx, column=7, value=email_data.get("website", ""))
    
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 25
    
    wb.save(output_path)
    return output_path
